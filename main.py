"""
2023-framework for multi-view4 sparse main process
"""

import time
import numpy as np
import scipy.io as io
import random
import pandas as pd
import xlsxwriter
import xlrd
from xlutils.copy import copy

import warnings
warnings.filterwarnings("ignore", category=FutureWarning, module="sklearn", lineno=70)

from skmultilearn.adapt import MLkNN
from sklearn.metrics import hamming_loss
from sklearn.metrics import zero_one_loss
from sklearn.metrics import label_ranking_loss
from sklearn.metrics import coverage_error
from sklearn.metrics import label_ranking_average_precision_score

import alg.DLHI as DLHI

def read_general_mat_data(filename):
    matr1 = io.loadmat(filename)
    matr1_list = list(matr1.keys())
    #print(matr1_list)
    para = {}
    for i in range(3, len(matr1_list)):
        para[matr1_list[i]] = matr1[matr1_list[i]]

    return para

def discrete_3(X):
    sample_row, sample_colum = np.shape(X)
    for i in range(sample_colum):
        m = pd.cut(X[:, i], 3, labels=[0, 1, 2])
        for j in range(sample_row):
            X[j, i] = m[j]
    return X

def get_data(f):
    matr1 = read_general_mat_data(f)
    ori_data = matr1['view']

    if 'features' not in matr1.keys():
        view_len = matr1['view'].shape[1]
        split_lst = []
        for s in range(view_len):
            split_lst.append(matr1['view'][0][s].shape[1])
        matr1['features'] = np.array(split_lst).reshape(view_len, 1)

    split_point = matr1['features']
    target = matr1['label']

    # integrate data
    ori_len = len(ori_data[0])
    c_data = ori_data[0][0]
    for i in range(1,ori_len):
        c_data = np.hstack((c_data, ori_data[0][i]))

    if f.split('\\')[-1] == 'emotions.mat' or f.split('\\')[-1] == 'yeast.mat':
        c_data = discrete_3(c_data)

    return c_data, split_point, target


def split_data_fold(X, Y, shuffle, n_test, i):
    m = Y.shape[0]
    if shuffle:
        random.shuffle(X)
    start_ind = i * n_test
    if (start_ind + n_test) > m:
        testI = np.arange(start_ind, m)
    else:
        testI = np.arange(start_ind, (start_ind + n_test))
    trainI = np.setdiff1d(np.arange(m), testI)

    trainX = X[trainI, :]
    Ytrain = Y[trainI, :]
    testX = X[testI, :]
    Ytest = Y[testI, :]
    return trainX, Ytrain, testX, Ytest

def read_data(X, F):
    """
    从X中获取F指定特征标号的数据存入data中
    """
    data=np.zeros((np.shape(X)[0],len(F)))
    z=0
    for i in F:
        i=int(float(i))
        data[:,z]=X[:,i]
        z=z+1
    return data

def writeperform(filename, dataname, Is, start_pos, flag, select_nub):
    alg = ["view4"]
    leng = len(alg)
    if flag == 1:  # 首次创建train表
        workbook = xlsxwriter.Workbook(filename)
        worksheet = workbook.add_worksheet()
        worksheet.write(start_pos - 1, 0, dataname)
        for i in range(leng):
            worksheet.write(start_pos + i, 0, alg[i])
        for i in range(leng):
            for j in range(1, select_nub + 1):
                worksheet.write(start_pos + i, j, float(Is[i][j - 1]))
        workbook.close()
    else:
        if dataname == "corel5k_5" or dataname == "iaprtc12_5_4999" or dataname == "3sources" or dataname == "espgame_4999" or dataname=="iaprtc12_5_9999" or dataname == "espgame_9999" or dataname == "nuswide_4999":
            import openpyxl
            data = openpyxl.load_workbook(filename)
            table = data['Sheet1']
            #rows = table.max_row
            temp = []
            temp.append([dataname])
            temp1 = []

            temp1.append(alg[0]) #只有一个算法
            for i in range(leng):
                for j in range(1, select_nub + 1):
                    temp1.append(float(Is[i][j - 1]))
            temp.append(temp1)
            for row_data in temp:
                table.append(row_data)
            data.save(filename)
        else:
            data = xlrd.open_workbook(filename)
            table = data.sheet_by_name('Sheet1')
            wb = copy(data)
            ws = wb.get_sheet(0)
            ws.write(start_pos - 1, 0, dataname)
            for i in range(leng):
                ws.write(start_pos + i, 0, alg[i])
            for i in range(leng):
                for j in range(1, select_nub + 1):
                    ws.write(start_pos + i, j, float(Is[i][j - 1]))
            wb.save(filename)

def writeperform_obj(filename, dataname, Is, start_pos, flag, select_nub):

    alg = ["view4"]
    leng = len(alg)
    if flag == 1:  # 首次创建train表
        workbook = xlsxwriter.Workbook(filename)
        worksheet = workbook.add_worksheet()
        worksheet.write(0,start_pos - 1, dataname)
        for i in range(leng):
            worksheet.write(0, start_pos + i, alg[i])
        for i in range(leng):
            for j in range(1, select_nub + 1):
                worksheet.write(j,start_pos + i, float(Is[i][j - 1]))
        workbook.close()
    else:
        data = xlrd.open_workbook(filename)
        table = data.sheet_by_name('Sheet1')
        wb = copy(data)
        ws = wb.get_sheet(0)
        ws.write(0, start_pos - 1, dataname)
        for i in range(leng):
            ws.write(0,start_pos + i, alg[i])
        for i in range(leng):
            for j in range(1, select_nub + 1):
                ws.write(j,start_pos + i, float(Is[i][j - 1]))
        wb.save(filename)


if __name__ == '__main__':
    start = time.localtime()

    datapos = 1
    dataposmax = 1
    flag = 1
    alg_nub = 1

    data_addr = [
                 '.\\data\\emotions.mat', '.\\data\\yeast.mat', '.\\data\\SCENE.mat',
                 '.\\data\\OBJECT.mat','.\\data\\VOC07.mat','.\\data\\MIRFlickr.mat',
               ]
    data_len = len(data_addr)
    svm_para = [0.01, 10, 0.1,
                1, 1, 1]
    for i in range(0,1):
        print(data_addr[i])
        X, split_point, Y = get_data(data_addr[i])

        CCCC = svm_para[i]

        dataname = (data_addr[i].split('\\')[-1]).split('.')[0]

        select_num = int(X.shape[1] * 0.2)
        HLscores = np.zeros((1, select_num))
        RLscores = np.zeros((1, select_num))
        CVscores = np.zeros((1, select_num))
        APscores = np.zeros((1, select_num))
        ZLscores = np.zeros((1, select_num))

        ratio = 0.2
        shuffle = False
        nfold = int(1 / ratio)
        m = Y.shape[0]
        n_test = round(m / nfold)

        f = 1
        for j in range(nfold):


            trainX, Ytrain, testX, Ytest = split_data_fold(X, Y, shuffle, n_test, j)

            zero_row = np.where(~(np.any(Ytrain, axis=1)))
            zero_col = np.where(~(np.any(Ytrain, axis=0)))
            one_row = np.where((np.all(Ytrain, axis=1)))
            one_col = np.where((np.all(Ytrain, axis=0)))

            Ytrain = np.delete(Ytrain, zero_row, axis=0)
            trainX = np.delete(trainX, zero_row, axis=0)
            Ytrain = np.delete(Ytrain, zero_col, axis=1)
            Ytest = np.delete(Ytest, zero_col, axis=1)

            # 特征排序:
            alpha_best = 1
            beta_best = 1
            gamma_best = 1
            lamb_best = 1
            record, iter = DLHI.DLHI(trainX,split_point,Ytrain,dataname,alpha_best,beta_best,gamma_best,lamb_best)

            # 评价
            print('evaluation:')

            select_num = int(X.shape[1]*0.2)

            top_20_lst = list(np.arange(1, select_num + 1))


            for k in range(1, len(top_20_lst) + 1):
                print("特征数目：", top_20_lst[k - 1])
                Fk = record['idx'][:top_20_lst[k - 1]]

                data_tr = read_data(trainX, Fk)
                data_te = read_data(testX, Fk)
                ########################################################################
                # MLkNN分类器
                classifier = MLkNN()
                classifier.fit(data_tr, Ytrain)
                predictions = classifier.predict(data_te)
                y_score = classifier.predict_proba(data_te)
                y_score = y_score.toarray()

                # 更新：
                HLscores[0][k - 1] = (HLscores[0][k - 1] * (j) + hamming_loss(Ytest, predictions)) / (j + 1)
                RLscores[0][k - 1] = (RLscores[0][k - 1] * j + label_ranking_loss(Ytest, y_score)) / (j + 1)
                CVscores[0][k - 1] = (CVscores[0][k - 1] * j + coverage_error(Ytest, y_score)) / (j + 1)
                APscores[0][k - 1] = (APscores[0][k - 1] * j + label_ranking_average_precision_score(Ytest,                                                                                        y_score)) / (j + 1)
                ZLscores[0][k - 1] = (ZLscores[0][k - 1] * j + zero_one_loss(Ytest, predictions)) / (j + 1)

            new_dataname = dataname + '_' + str(j)

        writeperform("result\\view4\\hamming_loss.xlsx", dataname, HLscores, datapos, flag,select_num)
        writeperform("result\\view4\\label_ranking_loss.xlsx", dataname, RLscores, datapos, flag, select_num)
        writeperform("result\\view4\\coverage_error.xlsx", dataname, CVscores, datapos, flag, select_num)
        writeperform("result\\view4\\average_precision_score.xlsx", dataname, APscores, datapos, flag, select_num)
        writeperform("result\\view4\\zero_one_loss.xlsx", dataname, ZLscores, datapos, flag, select_num)
        flag = 0
        datapos = datapos + 1 + alg_nub
        dataposmax = dataposmax + 3

    end = time.localtime()
    print("start time=", time.asctime(start))
    print(" end  time=", time.asctime(end))


